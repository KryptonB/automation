#########################################################################
#
# This script will log into ticketing website and scrape details of 
# delegated tickets and writes them to a report and sends it out
# to the specified recipients. 
#
#    Can extract delegated ticket details from ticketing website    
#    Creates an excel report with ticket details
#    Sends a mail with the report attached
#    Can be scheduled to run so that it monitors Inbox throughout
#    Can be customized to extract Opened / Pending / Reopened tickets
#    
# Developed by FSS Team
#########################################################################

import os
import sys
import time
import json
import logging
import datetime

import smtplib
from email import encoders
from email.mime.text import MIMEText 
from email.mime.base import MIMEBase 
from email.mime.multipart import MIMEMultipart

import openpyxl
from openpyxl.styles import colors
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Font, Color, PatternFill, Alignment

from selenium import webdriver
from selenium.common.exceptions import TimeoutException
from selenium.common.exceptions import WebDriverException
from selenium.common.exceptions import NoSuchElementException
from selenium.common.exceptions import UnexpectedAlertPresentException




# Wait intervals to be used later in the script until page is loaded in slow networks
tinyPause = 1
shortPause = 5
longPause = 10
longerPause = 20


# Set datestamp and timestamp to be used in file names
now = datetime.datetime.now()
datestamp = now.strftime('%Y%m%d_%H%M')
datestamp2 = now.strftime('%Y/%m/%d %H:%M')
timestamp = now.strftime('%Y%m%d%H%M%S')


# Global browser object
browser = ''


# Credentials for your ticketing website
username = ''
password = ''


# URL of your ticketing website
artoURL = ''


# Mail server host
mailServer = ''


# List to hold the delegated tickets
ticketList = []


# Ticket statuses to be included in the report
ticketStatuses = ['Delegated', 'Reopened', 'Pending', 'Opened']


# Ticket status colors - 87CEFA (blue), FFC0CB (pink)
statusColors = {'Delegated': 'FFDAB9', 'Reopened': 'AFEEEE', 'Pending': '98FB98', 'Opened': '87CEFA' }


# Ticket counts for each status. This will hold how many tickets in each status
ticketCounts = {}


# Total ticket count in the report
totalTicketCount = 0 #len(ticketList)


# User-wise ticket counts for each ticket status
openedCountForUser = {}
pendingCountForUser = {}
reopenedCountForUser = {}
delegatedCountForUser = {}


# Configurations file. This file holds all your personal configuration details
# such as website url, username, password, mail server host/ip
# If you don't like to keep them all in one file, you can create environment variables 
# for each config value and access them using the extract_configs_using_env_variables() function
configsFile = '.\\config\\configs.json'


# Chrome driver path
chromeDriverPath = '.\\drivers\\chromedriver.exe'


# Name of the report that will be generated. Make sure that the 'reports' folder exists. 
# If the 'reports' folder does not exist, you need to create it before running the script
reportName = '.\\reports\\Delegated_Ticket_List_' + datestamp + '.xlsx'


# Default log level. Only messages in and above this level will be printed to the log
# Accepted levels are DEBUG, INFO, WARNING, ERROR, CRITICAL
logLevel = logging.INFO


# Log file path and name. Make sure that the 'logs' folder exists. 
# If the 'logs' folder does not exist, you need to create it before running the script
logFile = '.\\logs\\ticket_report_' + datestamp + '.log'


# Log configurations
logFormat = '[%(asctime)s] %(levelname)s (%(filename)s: %(lineno)d) %(message)s'
logging.basicConfig(filename=logFile, level=logLevel, format=logFormat)


# Specify which columns should have which widths.
# Actual widths will be assigned later in the code
OBJECT_COLUMN = 'B'
TICKET_TITLE_COLUMN = 'D'
ASSIGNED_TO_COLUMN = 'G'
REGISTERED_FOR_COLUMN = 'E'
SECONDARY_STATUS_COLUMN = 'H'
MEDIUM_WIDTH_COLUMNS = ['C']
SMALL_WIDTH_COLUMNS = ['A', 'F']
LARGE_WIDTH_COLUMNS = ['D', 'E']


# Cells for report headings
TITLE_CELLS = ['A1', 'B1', 'C1', 'D1', 'E1', 'F1', 'G1']


# Format (bold) for report heading cells
TITLE_FONT = Font(bold=True)


# Styles for border of the tables in the 'SUMMARY' tab
thinBorder = Side(border_style='thin', color='000000')
OUTLINE_BORDER = Border(top=thinBorder, left=thinBorder, right=thinBorder, bottom=thinBorder)


# Create a new Excel workbook
wb = openpyxl.Workbook()


# Name for the summary sheet
summarySheetName = 'SUMMARY'
wb.get_active_sheet().title = summarySheetName




#########################################################################
#  Function to extract configuration details for this script
#########################################################################
def extract_configs(configsFile):
    global username
    global password
    global artoURL
    global mailServer
    
    try:
        with open(configsFile,'r') as f:
            config = json.load(f)
            
        username = config['user']['name']
        password = config['user']['password']
        logging.info('Username / Password extracted successfully')
        
        artoURL = config['website']['url']
        logging.info('Ticketing tool URL extracted successfully')
        
        mailServer = config['mailServer']['host']
        logging.info('Mail server host details extracted successfully')
            
    except IOError as e:
        # Do this if there is an error in the file path / folder
        print('File not found!\n' + str(e))
        logging.error(e, exc_info=True)
        logging.info('Terminating the program...')
        sys.exit()
        
    except Exception as e:
        # Do this if there is some other error
        print('Error in file!\n' + str(e))
        logging.error(e, exc_info=True)
        logging.info('Terminating the program...')
        sys.exit()




#########################################################################
#  Function to extract configuration details from environment variables
#########################################################################
def extract_configs_using_env_variables():
    global username
    global password
    global artoURL
    global mailServer
    
    # Get ticketing website URL from your environment variable
    artoURL = os.environ.get('ARTO_URL')

    # Get the ticketing tool website username and password from your environment variable
    username = os.environ.get('ARTO_USERNAME')
    password = os.environ.get('ARTO_PASSWORD')
    
    # Get mail server host from the environment variable
    mailServer = os.environ.get('MAIL_SERVER')




#########################################################################   
#  Utility function to check file paths and folders
#########################################################################
def check_file_and_folder_paths(filePath):

    # Get the file name from the given path
    givenFileName = os.path.basename(filePath)
    
    # Get the folder name from the given path
    givenFolderName = os.path.dirname(filePath)
    
    if not os.path.exists(givenFolderName):
        # Folder does not exist
        print('Folder: ' + givenFolderName + ' Does not exist!')
        logging.error('Folder: ' + givenFolderName + ' Does not exist!')
        return False
        
    elif not os.path.exists(filePath):
        # File does not exist
        print('File: ' + givenFileName + ' Does not exist!')
        logging.error('File: ' + givenFileName + ' Does not exist!')
        return False
        
    else:
        # Both file and folder exist. Therefore, no issue!
        return True

  


#########################################################################
#  Function to initiate browser  and start scraping
#########################################################################
def initiate_scraping(url, username, password):
    global browser
    global ticketList
    global totalTicketCount
    
    try:
        # Check the driver path and file
        if not (check_file_and_folder_paths(chromeDriverPath)):
            print(chromeDriverPath + ' Does not exist')
            logging.error(chromeDriverPath + ' Does not exist!')
            logging.info('Terminating the program...')
            sys.exit()
        else:
            # Initiate a new browser instance and load the ticketing website
            browser = webdriver.Chrome(chromeDriverPath)
            browser.get(url)
            logging.info('Chrome driver path: ' + chromeDriverPath)
            logging.info('Website URL: ' + url)
        
    except WebDriverException as e:
        print('Selenium exception!\n' + str(e))
        logging.error(e, exc_info=True)
        logging.info('Terminating the program...')
        sys.exit()
    
    time.sleep(shortPause)
    
    try:
        # Find the username textbox and type the username in to it
        u = browser.find_element_by_id('UserName')
        u.click()
        u.send_keys(username)
        logging.info('Username box found')
        
        time.sleep(tinyPause)
        
        # Find the password textbox and type the password in to it
        p = browser.find_element_by_id('PassWord')
        p.click()
        p.send_keys(password)
        logging.info('Password box found')
        
        # Find the login button and click it
        login = browser.find_element_by_xpath('/html/body/div/form/p[4]/input')
        login.click()
        logging.info('Login button found')

        # Now we are inside the site. Wait 20 seconds till the page is fully loaded 
        time.sleep(longerPause)

        # Find the left navigation pane's iframe and switch to that iframe
        leftNavigationFrame = browser.find_element_by_id('Page3')
        browser.switch_to_frame(leftNavigationFrame)
        logging.info('Switched to left navigation menu')
        
        # Find the "Tickets" menu button in the left pane and click it
        ticketsMenuButton = browser.find_element_by_id('td2')
        ticketsMenuButton.click()
        logging.info('Tickets button found and clicked!')

        time.sleep(longerPause)
        
        # Leave the left menu iframe and come back to the default frame
        browser.switch_to_default_content()

        # Find the top pane's iframe and switch to that iframe
        topFrame = browser.find_element_by_id('Set2')
        browser.switch_to_frame(topFrame)
        logging.info('Switched to top pane')
        
        # Find the dropdown list's frame and switch to that iframe
        menuFrame = browser.find_element_by_id('Page2')
        browser.switch_to_frame(menuFrame)
        logging.info('Switched to dropdown list pane')
        
        time.sleep(shortPause)

        # Find the dropdown list (select box) and click it
        tlist = browser.find_element_by_id('ListID')
        tlist.click()
        logging.info('Dropdown list found and clicked!')
        
        # Find the 'All new and Current Tickets' option from the dropdown list and click it
        allNew_and_Current = browser.find_element_by_xpath('//*[@id="ListID"]/option[3]')
        allNew_and_Current.click()
        logging.info('All New and Current Tickets option from the list is found and clicked!')
        
        # Click the dropdown list again to hide the list items
        tlist.click()

        # Wait until the open tickets are loaded. This will take some time
        time.sleep(longerPause)

        # Leave the top pane and come back to the default frame
        browser.switch_to_default_content()

        # Find the iframe where ticket details are displayed and switch to that frame
        newTicketFrame = browser.find_element_by_id('Page4')
        browser.switch_to_frame(newTicketFrame)
        logging.info('Switched to tickets pane')

        # Wait for 20 seconds until all ticket details are loaded
        time.sleep(longerPause)
        
        # Here is a sample xpath expression to find all 'Delegated' tickets
        #delegatedTicketRows = browser.find_elements_by_xpath('//*[@id="tblErrands"]/tbody/tr[contains(td[8], "Delegated")]')
        
        # Loop thru each ticket status and extract their details
        for ticketStatus in ticketStatuses:
        
            # Find all rows where the ticket status is the given status.
            ticketRows = browser.find_elements_by_xpath('//*[@id="tblErrands"]/tbody/tr[contains(td[8],' + ' "' + ticketStatus + '")]')
            logging.info('Reading ticket details...')
            
            # Number of tickets with the given status
            foundCount = len(ticketRows)
        
            if (foundCount != 0):
                # Since there are tickets under the given category, we write them to the report
                write_report(ticketStatus, ticketRows)
                
                # Add the ticket count to the ticketCounts dictionary 
                ticketCounts[ticketStatus] = foundCount
                
                print(str(foundCount) + ' ' + ticketStatus + ' tickets found' + '\n')
                logging.info(str(foundCount) + ' ' + ticketStatus + ' tickets found')
                
            else:
                # There are no tickets for the selected status. Hence no need to write anything to the report
                print('There are no ' + ticketStatus + ' tickets')
                logging.info('There are no ' + ticketStatus + ' tickets')
                
            # Assign the found tickets to the global variable so that it can be used later
            ticketList += ticketRows
        
        try:
            totalTicketCount = len(ticketList)
            
            # Call the method to create the summary sheet
            create_summary_sheet(wb, summarySheetName)
        
            # Save the report
            wb.save(reportName)
            logging.info('Saving the report...')
            logging.info('Saved successfully. Report name: ' + reportName)
            
            print('\n' + 'Report created successfully')
            
        except IOError as e:
            print('I/O Error! No such file or directory! Or permission issue! \n' + str(e))
            logging.error(e, exc_info=True)
            logging.info('Terminating the program...')
            browser.quit()
            sys.exit()
    
    except NoSuchElementException as e:
        print('Element not found!\n' + str(e))
        logging.error(e, exc_info=True)
        logging.info('Terminating the program...')
        browser.quit()
        sys.exit()
        
    except UnexpectedAlertPresentException as e:
        print('Username or Password is incorrect!\n' + str(e))
        logging.error(e, exc_info=True)
        logging.info('Terminating the program...')
        browser.quit()
        sys.exit()




#########################################################################   
#  Function for creating the excel report
#########################################################################
def write_report(ticketStatus, ticketRows):

    # Create a new sheet for each ticket status
    #sheet = wb.create_sheet(index=1, title=ticketStatus)
    sheet = wb.create_sheet(title=ticketStatus)

    # Rename the 1st sheet of the newly created workbook
    #sheet.title = 'Delegated Tickets'
    
    # Set the relevant color for the given ticket status. This color will be used 
    # for headings and tab color
    ticketStatusColor = statusColors[ticketStatus]
    
    # Change the tab color of the active sheet
    # 1072BA - blue, 
    sheet.sheet_properties.tabColor = ticketStatusColor
    
    # Background color for report column cells
    # Some more colors: feedc6 (light orange) / FFA07A (darker orange) / F0E68C (light green) / FFDAB9 (light orange)
    # If you want to remove the gradient, use fill_type='solid'. If you want a gradient, use fill_type=None
    BACKGROUND_FILL = PatternFill(start_color=ticketStatusColor, end_color=ticketStatusColor, fill_type='solid')

    # Add formatting to the heading cells
    for pos in TITLE_CELLS:
        sheet[pos].font = TITLE_FONT
        sheet[pos].fill = BACKGROUND_FILL

    # Center the headings
    for row_cells in sheet.iter_rows():
        for cell in row_cells:
            cell.alignment = Alignment(horizontal='center', vertical='center')

    # Set column width for small-width column
    for col in SMALL_WIDTH_COLUMNS:
        sheet.column_dimensions[col].width = 15

    # Set column width for medium-width column
    for col in MEDIUM_WIDTH_COLUMNS:
        sheet.column_dimensions[col].width = 20
    
    # Set column width for other remaining columns
    sheet.column_dimensions[OBJECT_COLUMN].width = 24
    sheet.column_dimensions[ASSIGNED_TO_COLUMN].width = 27
    sheet.column_dimensions[TICKET_TITLE_COLUMN].width = 40
    sheet.column_dimensions[REGISTERED_FOR_COLUMN].width = 40

    # Write column headings
    sheet['A1'] = 'Ticket number'
    sheet['B1'] = 'Object/Area'
    sheet['C1'] = 'Ticket type'
    sheet['D1'] = 'Ticket title'
    sheet['E1'] = 'Registered for'
    sheet['F1'] = 'Priority'
    sheet['G1'] = 'Assigned to / Status'
    
    # If it is a 'Pending' ticket, add an additional 'H' column to store the Secondary Status
    if(ticketStatus == 'Pending'):
        sheet['H1'] = 'Secondary Status'
        sheet.column_dimensions[SECONDARY_STATUS_COLUMN].width = 34
        sheet['H1'].font = TITLE_FONT
        sheet['H1'].fill = BACKGROUND_FILL

    # First row number which contains tickets details
    startingRow = 2
        
    # Loop through the ticket rows and print the details to the console
    for tableRow in ticketRows:
    
        # Find all <td> elements for the current <tr>
        cols = tableRow.find_elements_by_tag_name('td')
        
        # Print the values to the console. Just for testing purposes
        print('Ticket number: ' + cols[1].text)
        print('Object/Area: ' + cols[2].text)
        print('Ticket type: ' + cols[3].text)
        #print('Ticket title: ' + cols[4].text.encode('utf-8'))
        #print('Registered for: ' + cols[5].text)
        print('Priority: ' + cols[6].text)
        print('Assigned to: ' + cols[7].text)
        print('---------------------------------------------------------')
        
        logging.info('Ticket number: ' + cols[1].text)
        logging.info('Object/Area: ' + cols[2].text)
        logging.info('Ticket type: ' + cols[3].text)
        logging.info('Ticket title: ' + cols[4].text)
        logging.info('Registered for: ' + cols[5].text)
        logging.info('Priority: ' + cols[6].text)
        logging.info('Assigned to: ' + cols[7].text)

        # Write the details to the excel report
        sheet['A' + str(startingRow)] = cols[1].text
        sheet['A' + str(startingRow)].alignment = Alignment(horizontal='center')
        sheet['B' + str(startingRow)] = cols[2].text
        sheet['C' + str(startingRow)] = cols[3].text
        sheet['D' + str(startingRow)] = cols[4].text
        
        # Extract details from 'Registered for' column
        reg = cols[5].text
        reg2 = reg.split('\n')
        sheet['E' + str(startingRow)] = reg2[0] + ' - ' + reg2[1]
        # If you don't want to split the values, then use below line
        #sheet['E' + str(startingRow)] = cols[5].text
        
        sheet['F' + str(startingRow)] = cols[6].text
        sheet['F' + str(startingRow)].alignment = Alignment(horizontal='center')
        
        # Extract details from the 'Assigned to / status' column
        usr = cols[7].text
        usr2 = usr.split('\n')
        sheet['G' + str(startingRow)] = usr2[1]
        # Use below code if you want to slice the column and remove the 'Delegated' keyword
        #sheet['G' + str(startingRow)] = cols[7].text[9:]
        
        if(ticketStatus == 'Pending'):
            sheet['H' + str(startingRow)] = cols[8].text
            
        # Increment ticket count of the user
        # If name is not already in it, add it to the ticket count dictionary
        if(ticketStatus == 'Opened'):
            openedCountForUser[usr2[1]] = openedCountForUser.get(usr2[1], 0) + 1
        elif(ticketStatus == 'Pending'):
            pendingCountForUser[usr2[1]] = pendingCountForUser.get(usr2[1], 0) + 1
        elif(ticketStatus == 'Reopened'):
            reopenedCountForUser[usr2[1]] = reopenedCountForUser.get(usr2[1], 0) + 1
        elif(ticketStatus == 'Delegated'):
            delegatedCountForUser[usr2[1]] = delegatedCountForUser.get(usr2[1], 0) + 1

        logging.info('Data written to the report. Moving to the next row...')
        logging.info('---------------------------------------------------------')
        
        # Increment the row number by 1 so that the next ticket writes to next row
        startingRow += 1

    logging.info('Data writing completed for ' + ticketStatus + ' tickets')




#########################################################################    
# Function to create the SUMMARY sheet of the report
#########################################################################
def create_summary_sheet(workbook, sheetName):

    # Switch to the SUMMARY sheet
    sheet = workbook.get_sheet_by_name(sheetName)
    
    # Hide grid lines
    sheet.sheet_view.showGridLines = False

    # Column names and cell names for headings
    USER_COLUMNS = ['A', 'D', 'G', 'J']
    COUNT_COLUMNS = ['B', 'E', 'H', 'K']    
    USER_COLUMN_HEADING_CELLS = ['A2', 'D2', 'G2', 'J2']
    COUNT_COLUMNS_HEADING_CELLS = ['B2', 'E2', 'H2', 'K2']
    SUMMARY_TITLE_COLUMNS = ['A2', 'B2', 'D2', 'E2', 'G2', 'H2', 'J2', 'K2']
    
    MERGE_CELLS_LIST = ['A1:B1', 'D1:E1', 'G1:H1', 'J1:K1']
    SUMMARY_TITLE_MERGED_FIRST_CELLS = ['A1', 'D1', 'G1', 'J1']
    SUMMARY_TITLE_MERGED_SECOND_CELLS = ['B1', 'E1', 'H1', 'K1']
    
    # Format (bold) for report heading cells
    SUMMARY_TITLE_FONT = Font(bold=True, color='FFFFFF')
    
    # Format (bold) for merged heading cells
    SUMMARY_TITLE_MERGED_FONT = Font(bold=True)
    
    # Background color for summary title second row headings (user and count headings) A0522D - brown
    SUMMARY_TITLE_BACKGROUND_FILL = PatternFill(start_color='B22222', end_color='B22222', fill_type='solid')
    
    # Background color for summary title merged cells - FF8C00 - orange
    SUMMARY_TITLE_MERGED_CELL_BACKGROUND_FILL = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')
    
    # Set the tab color
    sheet.sheet_properties.tabColor = 'FFA500'
    
    for pos in SUMMARY_TITLE_COLUMNS:
        sheet[pos].font = SUMMARY_TITLE_FONT
        sheet[pos].fill = SUMMARY_TITLE_BACKGROUND_FILL
        sheet[pos].border = OUTLINE_BORDER
    
    # Add 'User' text to user columns
    for pos in USER_COLUMN_HEADING_CELLS:
        sheet[pos] = 'User'
    
    # Add 'Count' text to count columns
    for pos in COUNT_COLUMNS_HEADING_CELLS:
        sheet[pos] = 'Count'

    # Set column width for user column
    for pos in USER_COLUMNS:
        sheet.column_dimensions[pos].width = 27
        
    # Set column width for count column
    for pos in COUNT_COLUMNS:
        sheet.column_dimensions[pos].width = 8

    # Write merged cell headings
    sheet['A1'] = 'Opened Tickets Per User'
    sheet['D1'] = 'Delegated Tickets Per User'
    sheet['G1'] = 'Pending Tickets Per User'
    sheet['J1'] = 'Reopened Tickets'
      
    # Merge top row cells
    for pos in MERGE_CELLS_LIST:
        sheet.merge_cells(pos)
        
    # Set formatting for merged cells (first cell only)
    for pos in SUMMARY_TITLE_MERGED_FIRST_CELLS:
        sheet[pos].font = SUMMARY_TITLE_MERGED_FONT        
        sheet[pos].alignment = Alignment(horizontal='center')
        sheet[pos].fill = SUMMARY_TITLE_MERGED_CELL_BACKGROUND_FILL
        sheet[pos].border = OUTLINE_BORDER
        
    # Set border for second merged cell
    for pos in SUMMARY_TITLE_MERGED_SECOND_CELLS:
        sheet[pos].border = OUTLINE_BORDER
    
    # First row number which contains tickets details
    openedCounStartingRow = 3
    pendingCountStartingRow = 3
    reopenedCountStartingRow = 3
    delegatedCountStartingRow = 3
    
    if openedCountForUser:
        # Write open ticket user counts
        for k, v in openedCountForUser.items():
            sheet['A' + str(openedCounStartingRow)] = k
            sheet['A' + str(openedCounStartingRow)].border = OUTLINE_BORDER
            sheet['B' + str(openedCounStartingRow)] = v
            sheet['B' + str(openedCounStartingRow)].border = OUTLINE_BORDER
        
            # Increment row number by 1 so that the next ticket writes to next row
            openedCounStartingRow += 1       
    else:
        sheet['A' + str(openedCounStartingRow)] = 'No data'
            
    if delegatedCountForUser:
        # Write delegated ticket user counts
        for k, v in delegatedCountForUser.items():
            sheet['D' + str(delegatedCountStartingRow)] = k
            sheet['D' + str(delegatedCountStartingRow)].border = OUTLINE_BORDER
            sheet['E' + str(delegatedCountStartingRow)] = v
            sheet['E' + str(delegatedCountStartingRow)].border = OUTLINE_BORDER
        
            # Increment row number by 1 so that the next ticket writes to next row
            delegatedCountStartingRow += 1            
    else:
        sheet['D' + str(delegatedCountStartingRow)] = 'No data'
    
    if pendingCountForUser:
        # Write pending ticket user counts
        for k, v in pendingCountForUser.items():
            sheet['G' + str(pendingCountStartingRow)] = k
            sheet['G' + str(pendingCountStartingRow)].border = OUTLINE_BORDER
            sheet['H' + str(pendingCountStartingRow)] = v
            sheet['H' + str(pendingCountStartingRow)].border = OUTLINE_BORDER
        
            # Increment row number by 1 so that the next ticket writes to next row
            pendingCountStartingRow += 1            
    else:
        sheet['G' + str(pendingCountStartingRow)] = 'No data'
    
    if reopenedCountForUser:
        # Write reopened ticket user counts
        for k, v in reopenedCountForUser.items():
            sheet['J' + str(reopenedCountStartingRow)] = 'Reopened count'
            sheet['J' + str(reopenedCountStartingRow)].border = OUTLINE_BORDER
            sheet['K' + str(reopenedCountStartingRow)] = v
            sheet['K' + str(reopenedCountStartingRow)].border = OUTLINE_BORDER
        
            # Increment row number by 1 so that the next ticket writes to next row
            reopenedCountStartingRow += 1            
    else:
        sheet['J' + str(reopenedCountStartingRow)] = 'No data'




#########################################################################  
#  Function to send the email with report attachment
#########################################################################
def send_report_email(reportName):

    # Specify From and To addresses
    fromAddress = 'tickets-report@amazing.com'
    recipients = ['bgates@amazing.com', 'sjobs@amazing.com']
    
    # Instantiate MIMEMultipart obj
    msg = MIMEMultipart() 
      
    # Store sender's email address in the msg  
    msg['From'] = fromAddress 
      
    # Store receiver's email address in the msg 
    msg['To'] = ", ".join(recipients) 
      
    # Store the email subject in the msg 
    msg['Subject'] = 'Tickets Report - As at ' + datestamp2

    # Text file that holds the content for the mail body
    mailBodyContentFile = '.\\config\\content.txt'

    # Open and read the text file and store it as a string
    fp = open(mailBodyContentFile, 'rb')
    body = fp.read()
      
    # Attach the body to the msg  
    msg.attach(MIMEText(body, 'plain')) 
      
    # Open ticket report  
    filename = 'Delegated_Ticket_List_' + datestamp + '.xlsx'
    attachment = open(reportName, "rb") 
      
    # Instance of MIMEBase 
    payload = MIMEBase('application', 'octet-stream') 
      
    # Change payload into encoded form 
    payload.set_payload((attachment).read()) 
      
    # Encode into base64 
    encoders.encode_base64(payload) 
    
    payload.add_header('Content-Disposition', "attachment; filename= %s" % filename) 
      
    # Attach payload to msg 
    msg.attach(payload) 
      
    # Initiate SMTP session 
    session = smtplib.SMTP(mailServer)
      
    # Convert the Multipart msg into a string 
    text = msg.as_string() 
      
    # Send email 
    session.sendmail(fromAddress, recipients, text) 
      
    # Terminate the session 
    session.quit()


 

#########################################################################
#  Function to get the ticket count in each status
#########################################################################
def get_ticket_count(ticketCountDict):
    for k, v in ticketCountDict.items():
        print(k + ' - ' + str(v) + ' tickets')
        logging.info(k + ' - ' + str(v) + ' tickets')




# If you are extracting config details using the json file, you need to call the below function.
# But if you are extracting config details via the environment variables, you need to 
# comment below line and uncomment the call to extract_configs_using_env_variables()
# I'm going with env variables method       
#extract_configs(configsFile)
extract_configs_using_env_variables()

# Start the web scraping
initiate_scraping(artoURL, username, password)

time.sleep(shortPause)

# Check if there are any tickets. If yes, then write the counts to the console and log
# and email the report
if ticketList:
    print('There are total ' + str(totalTicketCount) + ' tickets in the report')
    logging.info('There are total ' + str(totalTicketCount) + ' tickets in the report')
    
    # Attach the report and email it. 
    # Comment this line if you don't want to send an email report
    #send_report_email(reportName)
 
else:
    # If you come to this block, it means no tickets to write
    print('There are no tickets. Therefore, no report will be created!')
    logging.info('There are no tickets. Therefore, no report will be created!')

# Close the browser process
browser.quit()

print('There are ' + str(totalTicketCount) + ' tickets in the report')
logging.info('There are ' + str(totalTicketCount) + ' tickets in the report')

# Find the ticket counts and write it to the console and log
get_ticket_count(ticketCounts)