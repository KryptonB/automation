import time
import datetime
import openpyxl
from openpyxl.styles import Font, Color, PatternFill, Alignment
from openpyxl.styles import colors
from openpyxl.styles.borders import Border, Side
from selenium import webdriver
from selenium.webdriver.chrome.options import Options

import smtplib
from email.mime.multipart import MIMEMultipart 
from email.mime.text import MIMEText 
from email.mime.base import MIMEBase 
from email import encoders 

# Chrome driver path
chromeDriverPath = 'C:\\Users\\sratnappuli\\Desktop\\delegated_tickets_script\\drivers\\chromedriver.exe'

# Artologik website URL
artologikURL = 'https://fss.virtusa.com'

# Credentials for Artologik ticketing website
username = 'sratnappuli'
password = 'Winter123'

# List to hold the delegated tickets
delegatedTicketList = []

browser = ''

# Wait intervals to be used later in the script until page is loaded in slow networks
tinyPause = 2
shortPause = 5
longPause = 10
longerPause = 20

# Set datestamp and timestamp to be used in file names
now = datetime.datetime.now()
datestamp = now.strftime('%Y%m%d_%H%M')
timestamp = now.strftime('%Y%m%d%H%M%S')

# Name of the report that will be generated
reportName = '.\\reports\\Delegated_Ticket_List_' + datestamp + '.xlsx'

# Cells for report headings
TITLE_CELLS = ['A1', 'B1', 'C1', 'D1', 'E1', 'F1', 'G1']

# Format (bold) for report heading cells
TITLE_FONT = Font(bold=True)

# Background color for report column cells
BACKGROUND_FILL = PatternFill(start_color='feedc6', end_color='feedc6', fill_type='solid')

# Specify which columns should have which widths
SMALL_WIDTH_COLUMNS = ['A', 'F']
MEDIUM_WIDTH_COLUMNS = ['C']
LARGE_WIDTH_COLUMNS = ['D', 'E']
TICKET_TITLE_COLUMN = 'D'
REGISTERED_FOR_COLUMN = 'E'
DELEGATED_TO_COLUMN = 'G'
OBJECT_COLUMN = 'B'


# Function to initiate browser
def initiate_artologik_login(url, username, password):
    global delegatedTicketList
    global browser
    
    WINDOW_SIZE = '1920,1080'
    
    # options = Options()
    # #options = webdriver.ChromeOptions()
    # #options.add_argument('--headless')
    # options.headless = True
    # options.add_argument('--window-size=%s' % WINDOW_SIZE)
    # browser = webdriver.Chrome(executable_path=chromeDriverPath, chrome_options=options)
    
    browser = webdriver.Chrome(chromeDriverPath)
    browser.get(url)
    time.sleep(shortPause)
    
    # Find the username textbox and type the username in to it
    browser.find_element_by_id('UserName').click()
    browser.find_element_by_id('UserName').send_keys(username)
    
    print(url)
    print(username)
    print(password)
    
    time.sleep(tinyPause)
    
    # Find the password textbox and type the password in to it
    p = browser.find_element_by_id('PassWord')
    p.click()
    p.send_keys(password)
    
    # Find the login button and click it
    login = browser.find_element_by_xpath('/html/body/div/form/p[4]/input')
    login.click()
    print(login)

    # Now we are inside the site. Wait 20 seconds till the page is fully loaded 
    time.sleep(longerPause)

    # Find the left navigation pane's iframe and switch to that iframe
    leftNavigationFrame = browser.find_element_by_id('Page3')
    browser.switch_to_frame(leftNavigationFrame)
    print(leftNavigationFrame)
    
    # Find the "Tickets" menu button in the left pane and click it
    ticketsMenuButton = browser.find_element_by_id('td2')
    ticketsMenuButton.click()
    print(ticketsMenuButton)

    time.sleep(longerPause)
    
    # Leave the left menu iframe and come back to the default frame
    browser.switch_to_default_content()

    # Find the top pane's iframe and switch to that iframe
    topFrame = browser.find_element_by_id('Set2')
    browser.switch_to_frame(topFrame)
    print(topFrame)
    
    # Find the dropdown list's frame and switch to that iframe
    menuFrame = browser.find_element_by_id('Page2')
    browser.switch_to_frame(menuFrame)
    print(menuFrame)
    
    time.sleep(shortPause)

    # Find the dropdown list (select box) and click it
    tlist = browser.find_element_by_id('ListID')
    tlist.click()
    print(tlist)
    
    # Find the 'All new and Current Tickets' option from the dropdown list and click it
    allNew_and_Current = browser.find_element_by_xpath('//*[@id="ListID"]/option[3]')
    allNew_and_Current.click()
    print(allNew_and_Current)
    
    # Click the dropdown list again to hide the list items
    tlist.click()

    # Wait until the open tickets are loaded. This will take some time
    time.sleep(longerPause)

    # Leave the top pane and come back to the default frame
    browser.switch_to_default_content()

    # Find the iframe where ticket details are displayed and switch to that frame
    newTicketFrame = browser.find_element_by_id('Page4')
    browser.switch_to_frame(newTicketFrame)
    print(newTicketFrame)

    # Wait for 20 seconds until all ticket details are loaded
    time.sleep(longerPause)
    
    # Find all rows where the ticket status is 'Delegated'.
    #delegatedTicketRows = browser.find_elements_by_xpath('//*[@id="tblErrands"]/tbody/tr[contains(td[8], "Delegated")]')
    delegatedTicketRows = browser.find_elements_by_xpath('//*[@id="tblErrands"]/tbody/tr[contains(td[8], "Pending")]')
    print(delegatedTicketRows)
    
    print('There are ' + str(len(delegatedTicketRows)) + ' Deletaged tickets' '\n')
    
    delegatedTicketList = delegatedTicketRows
    
    
   
   
# Function for creating the excel report 
def create_report(delegatedTicketList):
    
    # Create a new Excel workbook
    wb = openpyxl.Workbook()
    sheet = wb.get_active_sheet()

    # Rename the 1st sheet of the newly created workbook
    sheet.title = 'Delegated Tickets'
    
    # Change the tab color of the active sheet
    sheet.sheet_properties.tabColor = "1072BA"

    # Add formatting to the heading cells
    for pos in TITLE_CELLS:
        sheet[pos].font = TITLE_FONT
        sheet[pos].fill = BACKGROUND_FILL

    # Center headings
    for row_cells in sheet.iter_rows():
        for cell in row_cells:
            cell.alignment = Alignment(horizontal='center', vertical='center')

    # Set column width for small-width column
    for col in SMALL_WIDTH_COLUMNS:
        sheet.column_dimensions[col].width = 15

    # Set column width for medium-width column
    for col in MEDIUM_WIDTH_COLUMNS:
        sheet.column_dimensions[col].width = 20
    
    # Set column width for other remaining columns
    sheet.column_dimensions[TICKET_TITLE_COLUMN].width = 40
    sheet.column_dimensions[REGISTERED_FOR_COLUMN].width = 32
    sheet.column_dimensions[DELEGATED_TO_COLUMN].width = 22
    sheet.column_dimensions[OBJECT_COLUMN].width = 24

    # Write column headings
    sheet['A1'] = 'Ticket number'
    sheet['B1'] = 'Object/Area'
    sheet['C1'] = 'Ticket type'
    sheet['D1'] = 'Ticket title'
    sheet['E1'] = 'Registerd for'
    sheet['F1'] = 'Priority'
    sheet['G1'] = 'Delegated to'

    # First row number which contains delegated tickets details
    startingRow = 2
        
    # Loop through the ticket rows and print the details to the console
    for tableRow in delegatedTicketList:
        # Find all <td> elements for the current <tr>
        cols = tableRow.find_elements_by_tag_name('td')
        
        # Print the values to the console. Just for testing purposes
        print('Ticket number: ' + cols[1].text)
        print('Object/Area: ' + cols[2].text)
        print('Ticket type: ' + cols[3].text)
        print('Ticket title: ' + cols[4].text)
        print('Registerd for: ' + cols[5].text)
        print('Priority: ' + cols[6].text)
        print('Status: ' + cols[7].text)
        print('---------------------------------')

        # Write the details to the excel report
        sheet['A' + str(startingRow)] = cols[1].text
        sheet['B' + str(startingRow)] = cols[2].text
        sheet['C' + str(startingRow)] = cols[3].text
        sheet['D' + str(startingRow)] = cols[4].text
        
        # Extract details from 'Registered for' column
        reg = cols[5].text
        reg2 = reg.split('\n')
        sheet['E' + str(startingRow)] = reg2[0] + ' - ' + reg2[1]
        # If you don't want to split the values then use below line
        #sheet['E' + str(startingRow)] = cols[5].text
        
        sheet['F' + str(startingRow)] = cols[6].text
        sheet['G' + str(startingRow)] = cols[7].text[9:]

        # Increment the row number by 1 so that the next ticket writes to next row
        startingRow += 1

    # Save the report
    wb.save(reportName)
    
    print('Report created successfully')


initiate_artologik_login(artologikURL, username, password)
time.sleep(shortPause)

# Check if there are any delegated tickets. If yes, then start creating the report
if (len(delegatedTicketList) != 0):
    create_report(delegatedTicketList)
else:
    # There are no delegated tickets currently. Therefore, inform the user about it in console
    print('There are no delegated tickets. Therefore, no report will be created!')

browser.quit()