#########################################################################
#
# This script will log into Artologik and scrape details of 
# delegated tickets and writes them to a report and sends it out
# to the specified recipients. 
#
#    Can extract delegated ticket details from Artologik    
#    Creates an excel report with ticket details
#    Sends a mail with the report attached
#    Can be scheduled to run so that it monitors inbox throughout
#    Can be customized to extract Opened / Pending / Reopened tickets
#    
# Developed by FSS Team
#########################################################################

import os
import sys
import time
import json
import logging
import datetime

import smtplib
from email import encoders
from email.mime.text import MIMEText 
from email.mime.base import MIMEBase 
from email.mime.multipart import MIMEMultipart

import openpyxl
from openpyxl.styles import colors
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Font, Color, PatternFill, Alignment

from selenium import webdriver
from selenium.common.exceptions import TimeoutException
from selenium.common.exceptions import WebDriverException
from selenium.common.exceptions import NoSuchElementException
from selenium.common.exceptions import UnexpectedAlertPresentException




# Wait intervals to be used later in the script until page is loaded in slow networks
tinyPause = 1
shortPause = 5
longPause = 10
longerPause = 20

# Set datestamp and timestamp to be used in file names
now = datetime.datetime.now()
datestamp = now.strftime('%Y%m%d_%H%M')
datestamp2 = now.strftime('%Y/%m/%d %H:%M')
timestamp = now.strftime('%Y%m%d%H%M%S')

# Global browser object
browser = ''

# Credentials for Artologik ticketing website
username = ''
password = ''

# List to hold the delegated tickets
ticketList = []

# Ticket statuses
ticketStatuses = ["Delegated", "Reopened", "Pending"]

# Artologik website URL
artologikURL = 'https://fss.virtusa.com'

# Credential file
credentialFile = 'C:\\Users\\sratnappuli\\Desktop\\delegated_tickets_script\\config\\credentials.json'

# Chrome driver path
chromeDriverPath = 'C:\\Users\\sratnappuli\\Desktop\\delegated_tickets_script\\drivers\\chromedriver.exe'

# Name of the report that will be generated
reportName = 'C:\\Users\\sratnappuli\\Desktop\\delegated_tickets_script\\reports\\Delegated_Ticket_List_' + datestamp + '.xlsx'

# Default log level. Only messages in and above this level will be printed
# Accepted levels are DEBUG, INFO, WARNING, ERROR, CRITICAL
logLevel = logging.INFO

# Log file path and name
logFile = 'C:\\Users\\sratnappuli\\Desktop\\delegated_tickets_script\\logs\\ticket_report_' + datestamp + '.log'

# Log configurations
logFormat = '[%(asctime)s] %(levelname)s (%(filename)s: %(lineno)d) %(message)s'
logging.basicConfig(filename=logFile, level=logLevel, format=logFormat)



# Function to extract credentials for login to Artologik
def extract_credentials(credentialFile):
    global username
    global password
    
    try:
        with open(credentialFile,'r') as f:
            config = json.load(f)
            
    except IOError as e:
        # There is an error in the file path / folder
        print('File not found!\n' + str(e))
        logging.error(e, exc_info=True)
        logging.info('Terminating the program...')
        sys.exit()
        
    except Exception as e:
        # There is some other error
        print('Error in file!\n' + str(e))
        logging.error(e, exc_info=True)
        logging.info('Terminating the program...')
        sys.exit()
    
    username = config['user']['name']
    password = config['user']['password']
    logging.info('Username / Password extracted successfully')
    

# Utility function to check file paths and folders
def check_file_and_folder_paths(filePath):

    # Get the file name from the given path
    givenFileName = os.path.basename(filePath)
    
    # Get the folder name from the given path
    givenFolderName = os.path.dirname(filePath)
    
    if not os.path.exists(givenFolderName):
        # Folder does not exist
        print('Folder: ' + givenFolderName + ' Does not exist!')
        logging.error('Folder: ' + givenFolderName + ' Does not exist!')
        return False
        
    elif not os.path.exists(filePath):
        # File does not exist
        print('File: ' + givenFileName + ' Does not exist!')
        logging.error('File: ' + givenFileName + ' Does not exist!')
        return False
        
    else:
        # Both file and folder exist. Therefore, no issue!
        return True
    

# Function to initiate browser
def initiate_artologik_login(url, username, password):
    global ticketList
    global browser
    
    try:
        # Check the driver path and file
        if not (check_file_and_folder_paths(chromeDriverPath)):
            print(chromeDriverPath + ' Does not exist')
            logging.error(chromeDriverPath + ' Does not exist!')
            logging.info('Terminating the program...')
            sys.exit()
        else:
            # Initiate a new browser instance and load the artologik website
            browser = webdriver.Chrome(chromeDriverPath)
            browser.get(url)
            logging.info('Chrome driver path: ' + chromeDriverPath)
            logging.info('Artologik URL: ' + url)
        
    except WebDriverException as e:
        print('Selenium exception!\n' + str(e))
        logging.error(e, exc_info=True)
        logging.info('Terminating the program...')
        sys.exit()
    
    time.sleep(shortPause)
    
    try:
        # Find the username textbox and type the username in to it
        u = browser.find_element_by_id('UserName')
        u.click()
        u.send_keys(username)
        logging.info('Username box found')
        
        time.sleep(tinyPause)
        
        # Find the password textbox and type the password in to it
        p = browser.find_element_by_id('PassWord')
        p.click()
        p.send_keys(password)
        logging.info('Password box found')
        
        # Find the login button and click it
        login = browser.find_element_by_xpath('/html/body/div/form/p[4]/input')
        login.click()
        logging.info('Login button found')

        # Now we are inside the site. Wait 20 seconds till the page is fully loaded 
        time.sleep(longerPause)

        # Find the left navigation pane's iframe and switch to that iframe
        leftNavigationFrame = browser.find_element_by_id('Page3')
        browser.switch_to_frame(leftNavigationFrame)
        logging.info('Switched to left navigation menu')
        
        # Find the "Tickets" menu button in the left pane and click it
        ticketsMenuButton = browser.find_element_by_id('td2')
        ticketsMenuButton.click()
        logging.info('Tickets button found and clicked!')

        time.sleep(longerPause)
        
        # Leave the left menu iframe and come back to the default frame
        browser.switch_to_default_content()

        # Find the top pane's iframe and switch to that iframe
        topFrame = browser.find_element_by_id('Set2')
        browser.switch_to_frame(topFrame)
        logging.info('Switched to top pane')
        
        # Find the dropdown list's frame and switch to that iframe
        menuFrame = browser.find_element_by_id('Page2')
        browser.switch_to_frame(menuFrame)
        logging.info('Switched to dropdown list pane')
        
        time.sleep(shortPause)

        # Find the dropdown list (select box) and click it
        tlist = browser.find_element_by_id('ListID')
        tlist.click()
        logging.info('Dropdown list found and clicked!')
        
        # Find the 'All new and Current Tickets' option from the dropdown list and click it
        allNew_and_Current = browser.find_element_by_xpath('//*[@id="ListID"]/option[3]')
        allNew_and_Current.click()
        logging.info('All New and Current Tickets option from the list is found and clicked!')
        
        # Click the dropdown list again to hide the list items
        tlist.click()

        # Wait until the open tickets are loaded. This will take some time
        time.sleep(longerPause)

        # Leave the top pane and come back to the default frame
        browser.switch_to_default_content()

        # Find the iframe where ticket details are displayed and switch to that frame
        newTicketFrame = browser.find_element_by_id('Page4')
        browser.switch_to_frame(newTicketFrame)
        logging.info('Switched to tickets pane')

        # Wait for 20 seconds until all ticket details are loaded
        time.sleep(longerPause)
        
        # Find all rows where the ticket status is 'Delegated'.
        delegatedTicketRows = browser.find_elements_by_xpath('//*[@id="tblErrands"]/tbody/tr[contains(td[8], "Delegated")]')
        #delegatedTicketRows = browser.find_elements_by_xpath('//*[@id="tblErrands"]/tbody/tr[contains(td[8], "Reopened")]')
        #delegatedTicketRows = browser.find_elements_by_xpath('//*[@id="tblErrands"]/tbody/tr[contains(td[8], "Pending")]')
        
        logging.info('Reading ticket details...')
        
        print(str(len(delegatedTicketRows)) + ' Deletaged tickets found' + '\n')
        logging.info(str(len(delegatedTicketRows)) + ' Delegated tickets found')
        
        # Assign the found tickets to the global variable so that it can be used in create_report() function
        ticketList = delegatedTicketRows
        
    except NoSuchElementException as e:
        print('Element not found!\n' + str(e))
        logging.error(e, exc_info=True)
        logging.info('Terminating the program...')
        browser.quit()
        sys.exit()
        
    except UnexpectedAlertPresentException as e:
        print('Username or Password is incorrect!\n' + str(e))
        logging.error(e, exc_info=True)
        logging.info('Terminating the program...')
        browser.quit()
        sys.exit()
   

   
# Function for creating the excel report 
def create_report(ticketList):



    # Specify which columns should have which widths.
    # Actual widths will be assigned later in the code
    OBJECT_COLUMN = 'B'
    TICKET_TITLE_COLUMN = 'D'
    ASSIGNED_TO_COLUMN = 'G'
    REGISTERED_FOR_COLUMN = 'E'
    MEDIUM_WIDTH_COLUMNS = ['C']
    SMALL_WIDTH_COLUMNS = ['A', 'F']
    LARGE_WIDTH_COLUMNS = ['D', 'E']
    
    # Cells for report headings
    TITLE_CELLS = ['A1', 'B1', 'C1', 'D1', 'E1', 'F1', 'G1']

    # Format (bold) for report heading cells
    TITLE_FONT = Font(bold=True)

    # Background color for report column cells
    # Other colors: feedc6 (light orange) / FFA07A (darker orange) / F0E68C (light green) / FFDAB9 (light orange)
    # If you want to remove the gradient, use fill_type='solid'. If you want a gradient, use fill_type=None
    BACKGROUND_FILL = PatternFill(start_color='FFDAB9', end_color='FFDAB9', fill_type='solid')
    
    # Create a new Excel workbook
    wb = openpyxl.Workbook()
    
    
    
    sheet = wb.get_active_sheet()

    # Rename the 1st sheet of the newly created workbook
    sheet.title = 'Delegated Tickets'
    
    # Change the tab color of the active sheet
    # 1072BA - blue, 
    sheet.sheet_properties.tabColor = "FFA07A"

    # Add formatting to the heading cells
    for pos in TITLE_CELLS:
        sheet[pos].font = TITLE_FONT
        sheet[pos].fill = BACKGROUND_FILL

    # Center headings
    for row_cells in sheet.iter_rows():
        for cell in row_cells:
            cell.alignment = Alignment(horizontal='center', vertical='center')

    # Set column width for small-width column
    for col in SMALL_WIDTH_COLUMNS:
        sheet.column_dimensions[col].width = 15

    # Set column width for medium-width column
    for col in MEDIUM_WIDTH_COLUMNS:
        sheet.column_dimensions[col].width = 20
    
    # Set column width for other remaining columns
    sheet.column_dimensions[OBJECT_COLUMN].width = 24
    sheet.column_dimensions[ASSIGNED_TO_COLUMN].width = 27
    sheet.column_dimensions[TICKET_TITLE_COLUMN].width = 40
    sheet.column_dimensions[REGISTERED_FOR_COLUMN].width = 40

    # Write column headings
    sheet['A1'] = 'Ticket number'
    sheet['B1'] = 'Object/Area'
    sheet['C1'] = 'Ticket type'
    sheet['D1'] = 'Ticket title'
    sheet['E1'] = 'Registered for'
    sheet['F1'] = 'Priority'
    sheet['G1'] = 'Assigned to / Status'

    # First row number which contains tickets details
    startingRow = 2
        
    # Loop through the ticket rows and print the details to the console
    for tableRow in ticketList:
    
        # Find all <td> elements for the current <tr>
        cols = tableRow.find_elements_by_tag_name('td')
        
        # Print the values to the console. Just for testing purposes
        print('Ticket number: ' + cols[1].text)
        print('Object/Area: ' + cols[2].text)
        print('Ticket type: ' + cols[3].text)
        print('Ticket title: ' + cols[4].text)
        print('Registered for: ' + cols[5].text)
        print('Priority: ' + cols[6].text)
        print('Assigned to: ' + cols[7].text)
        print('---------------------------------')
        
        logging.info('Ticket number: ' + cols[1].text)
        logging.info('Object/Area: ' + cols[2].text)
        logging.info('Ticket type: ' + cols[3].text)
        logging.info('Ticket title: ' + cols[4].text)
        logging.info('Registered for: ' + cols[5].text)
        logging.info('Priority: ' + cols[6].text)
        logging.info('Assigned to: ' + cols[7].text)

        # Write the details to the excel report
        sheet['A' + str(startingRow)] = cols[1].text
        sheet['A' + str(startingRow)].alignment = Alignment(horizontal='center')
        sheet['B' + str(startingRow)] = cols[2].text
        sheet['C' + str(startingRow)] = cols[3].text
        sheet['D' + str(startingRow)] = cols[4].text
        
        # Extract details from 'Registered for' column
        reg = cols[5].text
        reg2 = reg.split('\n')
        sheet['E' + str(startingRow)] = reg2[0] + ' - ' + reg2[1]
        # If you don't want to split the values, then use below line
        #sheet['E' + str(startingRow)] = cols[5].text
        
        sheet['F' + str(startingRow)] = cols[6].text
        sheet['F' + str(startingRow)].alignment = Alignment(horizontal='center')
        
        # Extract details from the 'Assigned to / status' column
        usr = cols[7].text
        usr2 = usr.split('\n')
        sheet['G' + str(startingRow)] = usr2[1]
        # Use below code to if you want to slice the column and remove the 'Delegated keyword'
        #sheet['G' + str(startingRow)] = cols[7].text[9:]

        logging.info('Data written to the report. Moving to the next row...')
        logging.info('---------------------------------')
        
        # Increment the row number by 1 so that the next ticket writes to next row
        startingRow += 1

    logging.info('Data writing completed')
    try:
        # Save the report
        wb.save(reportName)
        logging.info('Saving the report...')
        logging.info('Saved successfully. Report name: ' + reportName)
        
        print('\n' + 'Report created successfully')
        
    except IOError as e:
        print('I/O Error! No such file or directory! Or permission issue! \n' + str(e))
        logging.error(e, exc_info=True)
        logging.info('Terminating the program...')
        browser.quit()
        sys.exit()

    
# Function to send the email with report attachment
def send_report_email(reportName):

    # Specify From / To addresses
    fromAddress = 'delegated-tickets@virtusa.com'
    recipients = ['sratnappuli@virtusa.com', 'sratnappuli@virtusa.com']
    
    # Instantiate MIMEMultipart obj
    msg = MIMEMultipart() 
      
    # Store sender's email address in the msg  
    msg['From'] = fromAddress 
      
    # Store receiver's email address in the msg 
    msg['To'] = ", ".join(recipients) 
      
    # Store the email subject in the msg 
    msg['Subject'] = 'Delegated Tickets Report - As at ' + datestamp2

    # Text file that holds the content for the mail body
    mailBodyContentFile = 'C:\\Users\\sratnappuli\\Desktop\\delegated_tickets_script\\config\\content.txt'

    # Open and read the text file and store it as a string
    fp = open(mailBodyContentFile, 'rb')
    body = fp.read()
      
    # Attach the body to the msg  
    msg.attach(MIMEText(body, 'plain')) 
      
    # Open ticket report  
    filename = 'Delegated_Ticket_List_' + datestamp + '.xlsx'
    attachment = open(reportName, "rb") 
      
    # Instance of MIMEBase 
    payload = MIMEBase('application', 'octet-stream') 
      
    # Change payload into encoded form 
    payload.set_payload((attachment).read()) 
      
    # Encode into base64 
    encoders.encode_base64(payload) 
    
    payload.add_header('Content-Disposition', "attachment; filename= %s" % filename) 
      
    # Attach payload to msg 
    msg.attach(payload) 
      
    # Initiate SMTP session 
    session = smtplib.SMTP('10.62.65.37')
      
    # Convert the Multipart msg into a string 
    text = msg.as_string() 
      
    # Send email 
    session.sendmail(fromAddress, recipients, text) 
      
    # Terminate the session 
    session.quit()

    
    
    
extract_credentials(credentialFile)
initiate_artologik_login(artologikURL, username, password)
time.sleep(shortPause)

# Check if there are any delegated tickets. If yes, then start creating the report
if (len(ticketList) != 0):
    create_report(ticketList)
    #send_report_email(reportName)
else:
    # If you come to this block, it means no delegated tickets.
    print('There are no delegated tickets. Therefore, no report will be created!')

# Close the chrome browser process
browser.quit()